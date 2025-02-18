from io import BytesIO
from flask import request, make_response
from openpyxl import Workbook
from flask_shop.sku import sku_api
from flask_restful import Resource, reqparse
from flask_shop import models, db
import time
import random
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

class SKUs(Resource):
    def get(self):
        """获取SKU列表"""
        try:
            parser = reqparse.RequestParser()
            parser.add_argument('product_id', type=int, location='args')
            args = parser.parse_args()
            
            if args.get('product_id'):
                skus = models.SKU.query.filter_by(product_id=args.get('product_id')).all()
            else:
                skus = models.SKU.query.all()
                
            return {
                'status': 200,
                'msg': '获取SKU列表成功',
                'data': {
                    'total': len(skus),
                    'data': [sku.to_dict() for sku in skus]
                }
            }
        except Exception as e:
            print(f"获取SKU列表错误: {str(e)}")
            return {'status': 500, 'msg': str(e)}

    def post(self):
        """添加SKU"""
        try:
            # 获取请求数据
            name = request.get_json().get('product_id')
            specifications = request.get_json().get('specifications')
            price = request.get_json().get('price')
            stock = request.get_json().get('stock')
            
            # 验证必要字段
            if not all([name, specifications, price, stock]):
                return {'status': 400, 'msg': 'SKU参数不完整'}
            
            # 创建SKU
            sku = models.SKU(
                product_id=name,
                sku_code=self.generate_sku_code(),
                specifications=specifications,
                price=price,
                stock=stock,
                status=1  # 默认上架状态
            )
            
            # 保存到数据库
            db.session.add(sku)
            db.session.commit()
            return {'status': 200, 'msg': '添加SKU成功'}
        except Exception as e:
            db.session.rollback()
            return {'status': 500, 'msg': '添加SKU失败'}

    def generate_sku_code(self):
        timestamp = int(time.time())
        random_num = random.randint(1000, 9999)
        return f"SKU-{timestamp}-{random_num}"

sku_api.add_resource(SKUs, '/skus/')

class SKUManage(Resource):
    def put(self, id):
        parser = reqparse.RequestParser()
        parser.add_argument('price', type=float)
        parser.add_argument('stock', type=int)
        parser.add_argument('status', type=int)
        args = parser.parse_args()

        try:
            sku = models.SKU.query.get(id)
            if not sku:
                return {'status': 404, 'msg': 'SKU不存在'}

            if args.get('price'):
                sku.price = args.get('price')
            if args.get('stock') is not None:
                old_stock = sku.stock
                new_stock = args.get('stock')
                sku.stock = new_stock
                
                stock_log = models.StockLog(
                    sku_id=id,
                    change_amount=new_stock - old_stock,
                    type='manual',
                    operator=request.headers.get('username', 'unknown')
                )
                db.session.add(stock_log)
                
            if args.get('status') is not None:
                sku.status = args.get('status')
                
            db.session.commit()
            return {'status': 200, 'msg': '更新SKU成功'}
        except Exception as e:
            db.session.rollback()
            return {'status': 500, 'msg': str(e)}

    def delete(self, id):
        try:
            sku = models.SKU.query.get(id)
            if not sku:
                return {'status': 404, 'msg': 'SKU不存在'}
                
            db.session.delete(sku)
            db.session.commit()
            return {'status': 200, 'msg': '删除SKU成功'}
        except Exception as e:
            db.session.rollback()
            return {'status': 500, 'msg': '删除SKU失败'}

sku_api.add_resource(SKUManage, '/sku/<int:id>/')

class SKUExport(Resource):
    def get(self):
        """导出SKU列表为Excel"""
        try:
            # 获取所有SKU数据
            skus = models.SKU.query.all()
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "SKU列表"
            
            # 设置列宽
            column_widths = {
                'A': 15,  # SKU编码
                'B': 10,  # 商品ID
                'C': 30,  # 规格
                'D': 12,  # 价格
                'E': 10,  # 库存
                'F': 10,  # 销量
                'G': 10,  # 状态
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 定义样式
            header_fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type='solid')
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            data_font = Font(name='微软雅黑', size=10)
            
            # 边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 对齐方式
            center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_aligned = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # 写入表头
            headers = ['SKU编码', '商品ID', '规格', '价格', '库存', '销量', '状态']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_aligned
                cell.border = thin_border
            
            # 写入数据
            for row, sku in enumerate(skus, 2):
                # 格式化规格信息
                specs_str = '\n'.join([f"{k}: {v}" for k, v in sku.specifications.items()])
                
                row_data = [
                    sku.sku_code,
                    sku.product_id,
                    specs_str,
                    f"¥{sku.price:.2f}",
                    sku.stock,
                    sku.sales,
                    '上架' if sku.status == 1 else '下架'
                ]
                
                # 设置每个单元格的样式
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    # 规格列左对齐，其他居中
                    if col == 3:  # 规格列
                        cell.alignment = left_aligned
                    else:
                        cell.alignment = center_aligned
                    
                    # 设置交替行颜色
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            
            # 保存到内存
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            # 准备响应
            response = make_response(excel_file.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = 'attachment; filename=sku_list.xlsx'
            response.headers['Access-Control-Allow-Origin'] = '*'
            response.headers['Access-Control-Allow-Headers'] = 'Content-Type,token'
            response.headers['Access-Control-Allow-Methods'] = 'GET,POST,PUT,DELETE,OPTIONS'
            
            return response
            
        except Exception as e:
            print(f"Excel导出错误: {str(e)}")
            return {'status': 500, 'msg': '导出失败'}

sku_api.add_resource(SKUExport, '/skus/export/excel/') 